#!/usr/bin/env python3
# fpe_xlsx_tool.py
#
# Encrypt/decrypt selected columns in XLSX using Format-Preserving Encryption (FF1) with a manifest.
#
# ✅ Includes:
#   1) Stable tweak via row-id column (robust to sorting/reordering) OR row index (legacy).
#   2) Manifest integrity via HMAC (key derived from file_secret) + encrypted file SHA-256 binding.
#
# Engines:
#   - ff1-cloudproof-fpe (default): NIST SP 800-38G FF1 via `cloudproof-fpe` (Rust-backed wheel).
#   - demo-feistel-hmac-sha256     : legacy demo engine (NOT for production).
#
# Dependencies:
#   pip install openpyxl cloudproof-fpe
#
# Usage:
#   python fpe_xlsx_tool.py encrypt --in input.xlsx --out encrypted.xlsx --manifest manifest.json --config config.json --secret-out secret.b64
#   python fpe_xlsx_tool.py decrypt --in encrypted.xlsx --out decrypted.xlsx --manifest manifest.json --secret-b64 <...>
#
from __future__ import annotations

import argparse
import base64
import hashlib
import hmac
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

try:
    from cloudproof_fpe import Alphabet as CloudproofAlphabet  # pip install cloudproof-fpe
except Exception:  # pragma: no cover
    CloudproofAlphabet = None


# ----------------------------
# Helpers: base64, hashing, json canonicalization
# ----------------------------
_WS_RE = re.compile(r"\s+", flags=re.UNICODE)

def b64e(b: bytes) -> str:
    return base64.urlsafe_b64encode(b).decode("ascii")

def b64d(s: str) -> bytes:
    return base64.urlsafe_b64decode(s.encode("ascii"))

def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def canonical_json_bytes(obj: Any) -> bytes:
    # Stable serialization for HMAC
    return json.dumps(obj, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")

def norm_text(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = unicodedata.normalize("NFC", s)
    s = s.replace("\u00a0", " ")  # NBSP
    s = s.strip()
    s = _WS_RE.sub(" ", s)
    return s

def hkdf_like(ikm: bytes, salt: bytes, info: bytes, length: int) -> bytes:
    """HKDF (extract+expand) with HMAC-SHA256, minimal implementation."""
    prk = hmac.new(salt, ikm, hashlib.sha256).digest()
    okm = b""
    t = b""
    counter = 1
    while len(okm) < length:
        t = hmac.new(prk, t + info + bytes([counter]), hashlib.sha256).digest()
        okm += t
        counter += 1
    return okm[:length]


# ----------------------------
# Character classes & alphabets
# ----------------------------
ALPH_DIGITS = "0123456789"
ALPH_EN = "abcdefghijklmnopqrstuvwxyz"
ALPH_RU = "абвгдеёжзийклмнопрстуфхцчшщъыьэюя"  # 33 letters including ё

def is_cyrillic_char(ch: str) -> bool:
    o = ord(ch)
    return (0x0400 <= o <= 0x04FF) or (0x0500 <= o <= 0x052F) or (0x2DE0 <= o <= 0x2DFF) or (0xA640 <= o <= 0xA69F)

def apply_case_mask(transformed_lower: str, original: str) -> str:
    return "".join(ch.upper() if orig.isupper() else ch for ch, orig in zip(transformed_lower, original))


# ----------------------------
# Engine interface
# ----------------------------
class Engine:
    name: str = "engine"

    def transform_cell_value(self, value: Any, mode: str, key: bytes, tweak: bytes) -> Any:
        raise NotImplementedError


# ----------------------------
# Engine #1 (default): FF1 via cloudproof-fpe
# ----------------------------
class CloudproofFF1Engine(Engine):
    """
    Uses `cloudproof-fpe` (FF1) for digits and letters, while preserving:
      - cell length
      - digits -> digits
      - letters -> letters
      - case pattern (upper/lower) per character

    Note: For Cyrillic letters we extend the Latin alpha_lower alphabet with Russian letters,
    so ciphertext letters may become Latin or Cyrillic letters (still letters).
    """
    name = "ff1-cloudproof-fpe"

    def __init__(self, short_policy: str = "leave"):
        if CloudproofAlphabet is None:
            raise RuntimeError(
                "cloudproof_fpe is not available. Install with: pip install cloudproof-fpe"
            )
        if short_policy not in ("leave", "error"):
            raise ValueError("short_policy must be 'leave' or 'error'")
        self.short_policy = short_policy

        # FF1 alphabets
        self._a_digits = CloudproofAlphabet("numeric")     # 0-9
        self._a_en = CloudproofAlphabet("alpha_lower")     # a-z

        # For Cyrillic: start from a-z and add Russian lowercase letters.
        # This guarantees letters stay letters (no digits/symbols).
        self._a_ru = CloudproofAlphabet("alpha_lower")
        self._a_ru.extend_with(ALPH_RU)

    @staticmethod
    def _subkey(col_key: bytes, label: str, out_len: int = 32) -> bytes:
        return hkdf_like(col_key, salt=b"ff1-subkey", info=label.encode("utf-8"), length=out_len)

    @staticmethod
    def _subtweak(tweak: bytes, label: str, out_len: int = 16) -> bytes:
        return hkdf_like(tweak, salt=b"ff1-subtweak", info=label.encode("utf-8"), length=out_len)

    def _try_ff1(self, alphabet_obj, mode: str, key: bytes, tweak: bytes, text: str, what: str) -> str:
        try:
            if mode == "encrypt":
                return alphabet_obj.encrypt(key, tweak, text)
            return alphabet_obj.decrypt(key, tweak, text)
        except Exception as e:
            if self.short_policy == "leave":
                return text
            raise ValueError(f"FF1 {mode} failed for {what}: {e}") from e

    def _transform_digits(self, s: str, mode: str, key: bytes, tweak: bytes) -> str:
        if not any(ch.isdigit() for ch in s):
            return s
        k = self._subkey(key, "digits")
        tw = self._subtweak(tweak, "digits")
        # cloudproof-fpe encrypts only chars in alphabet and preserves others (e.g., '-' stays)
        return self._try_ff1(self._a_digits, mode, k, tw, s, what="digits")

    def _transform_letters(self, s: str, mode: str, key: bytes, tweak: bytes) -> str:
        # Collect letter positions. Cells may contain both Cyrillic and Latin letters (e.g. visually similar 'M' vs 'М').
        positions: List[int] = [i for i, ch in enumerate(s) if ch.isalpha()]
        if not positions:
            return s

        ru_positions = [i for i in positions if is_cyrillic_char(s[i])]
        en_positions = [i for i in positions if not is_cyrillic_char(s[i])]

        buf = list(s)

        # Transform Cyrillic letters (if any) using RU alphabet
        if ru_positions:
            label = "letters|ru"
            k = self._subkey(key, label)
            tw = self._subtweak(tweak, label)

            original = "".join(s[i] for i in ru_positions)
            lower = original.lower()
            out_lower = self._try_ff1(self._a_ru, mode, k, tw, lower, what=label)
            out = apply_case_mask(out_lower, original)

            for idx, pos in enumerate(ru_positions):
                buf[pos] = out[idx]

        # Transform Latin (and other non-Cyrillic) letters (if any) using EN alphabet
        if en_positions:
            label = "letters|en"
            k = self._subkey(key, label)
            tw = self._subtweak(tweak, label)

            original = "".join(s[i] for i in en_positions)
            lower = original.lower()
            out_lower = self._try_ff1(self._a_en, mode, k, tw, lower, what=label)
            out = apply_case_mask(out_lower, original)

            for idx, pos in enumerate(en_positions):
                buf[pos] = out[idx]

        return "".join(buf)

    def transform_cell_value(self, value: Any, mode: str, key: bytes, tweak: bytes) -> Any:
        # Keep formulas untouched
        if isinstance(value, str) and value.startswith("="):
            return value

        s = unicodedata.normalize("NFC", str(value))
        if not s:
            return s

        # IMPORTANT: use reverse order on decrypt (for safety, though digits/letters are disjoint)
        if mode == "encrypt":
            s = self._transform_digits(s, "encrypt", key, tweak)
            s = self._transform_letters(s, "encrypt", key, tweak)
        else:
            s = self._transform_letters(s, "decrypt", key, tweak)
            s = self._transform_digits(s, "decrypt", key, tweak)
        return s


class CloudproofFF1AlnumEngine(CloudproofFF1Engine):
    """
    Variant C (alphanumeric mix):

    Encrypt ALL characters that belong to ONE large alphabet (digits + Latin + Cyrillic, both cases).
    Non-alphabet characters (e.g. '-', '/', ' ') stay in place (cloudproof-fpe behavior).

    ✅ Pros:
      - fixes "short chunk" issues when values contain mixed letters+digits separated by punctuation,
        because the whole alphanumeric content participates as a single domain
      - no error on mixed Cyrillic+Latin inside one cell

    ⚠️ Trade-off:
      - digits may become letters and letters may become digits (still length-preserving)
      - case may change (because alphabet includes both a-z and A-Z, plus Cyrillic cases)
    """
    name = "ff1-cloudproof-fpe-alnum"

    def __init__(self, short_policy: str = "leave"):
        if CloudproofAlphabet is None:
            raise RuntimeError(
                "cloudproof_fpe is not available. Install with: pip install cloudproof-fpe"
            )
        # Do NOT call super().__init__ to avoid building unused alphabets;
        # but we reuse its helpers via direct attribute assignment.
        self.short_policy = short_policy

        # 0-9 a-z A-Z
        self._a_alnum = CloudproofAlphabet("alpha_numeric")

        # Extend with Cyrillic (both cases)
        ru_lower = ALPH_RU
        ru_upper = ru_lower.upper()
        self._a_alnum.extend_with(ru_lower + ru_upper)

    def transform_cell_value(self, value: Any, mode: str, key: bytes, tweak: bytes) -> Any:
        # Keep formulas untouched
        if isinstance(value, str) and value.startswith("="):
            return value

        s = unicodedata.normalize("NFC", str(value))
        if not s:
            return s

        label = "alnum"
        k = CloudproofFF1Engine._subkey(key, label)
        tw = CloudproofFF1Engine._subtweak(tweak, label)
        return CloudproofFF1Engine._try_ff1(self, self._a_alnum, mode, k, tw, s, what=label)


# ----------------------------
# Engine #2: Demo Feistel (legacy, NOT production)
# ----------------------------
def digits_to_int(digs: List[int], radix: int) -> int:
    x = 0
    for d in digs:
        x = x * radix + d
    return x

def int_to_digits(x: int, radix: int, length: int) -> List[int]:
    digs = [0] * length
    for i in range(length - 1, -1, -1):
        digs[i] = x % radix
        x //= radix
    return digs

def prf_int(key: bytes, tweak: bytes, round_no: int, r_digits: List[int], radix: int, out_len: int) -> int:
    r_bytes = bytes(r_digits)
    msg = tweak + b"|" + round_no.to_bytes(2, "big") + b"|" + bytes([radix]) + b"|" + r_bytes
    mac = hmac.new(key, msg, hashlib.sha256).digest()
    val = int.from_bytes(mac, "big")
    return val % (radix ** out_len)

def feistel_encrypt(digits: List[int], radix: int, key: bytes, tweak: bytes, rounds: int) -> List[int]:
    n = len(digits)
    if n < 2:
        return digits[:]
    nL = n // 2
    L = digits[:nL]
    R = digits[nL:]
    for i in range(rounds):
        m = len(L)
        F = prf_int(key, tweak, i, R, radix, m)
        L_int = digits_to_int(L, radix)
        new_L = R
        new_R = int_to_digits((L_int + F) % (radix ** m), radix, m)
        L, R = new_L, new_R
    return L + R

def feistel_decrypt(digits: List[int], radix: int, key: bytes, tweak: bytes, rounds: int) -> List[int]:
    n = len(digits)
    if n < 2:
        return digits[:]
    nL = n // 2
    nR = n - nL
    if rounds % 2 == 0:
        L = digits[:nL]
        R = digits[nL:]
    else:
        L = digits[:nR]
        R = digits[nR:]
    for i in range(rounds - 1, -1, -1):
        m = len(R)
        R_old = L
        F = prf_int(key, tweak, i, R_old, radix, m)
        R_int = digits_to_int(R, radix)
        L_old = int_to_digits((R_int - F) % (radix ** m), radix, m)
        L, R = L_old, R_old
    return L + R

def min_len_for_domain(radix: int, min_domain: int = 1_000_000) -> int:
    L = 1
    x = radix
    while x < min_domain:
        L += 1
        x *= radix
    return L

class DemoEngine(Engine):
    name = "demo-feistel-hmac-sha256"

    def __init__(self, rounds: int = 10, short_policy: str = "leave"):
        self.rounds = rounds
        self.short_policy = short_policy  # leave|error

    def transform_token(self, token: str, mode: str, key: bytes, tweak: bytes, alphabet: str) -> str:
        token_n = unicodedata.normalize("NFC", token)
        lower = token_n.lower()
        alpha_index = {ch: i for i, ch in enumerate(alphabet)}
        digs: List[int] = []
        for ch in lower:
            if ch not in alpha_index:
                return token_n
            digs.append(alpha_index[ch])
        radix = len(alphabet)

        min_len = min_len_for_domain(radix)
        if len(digs) < min_len:
            if self.short_policy == "error":
                raise ValueError(f"Token too short for radix={radix}: len={len(digs)} < min_len={min_len}: {token!r}")
            return token_n

        out = feistel_encrypt(digs, radix, key, tweak, self.rounds) if mode == "encrypt" else feistel_decrypt(digs, radix, key, tweak, self.rounds)
        out_lower = "".join(alphabet[d] for d in out)
        return apply_case_mask(out_lower, token_n)

    def transform_cell_value(self, value: Any, mode: str, key: bytes, tweak: bytes) -> Any:
        if isinstance(value, str) and value.startswith("="):
            return value

        s = unicodedata.normalize("NFC", str(value))
        if not s:
            return s

        # Tokenize into contiguous runs; demo only
        parts: List[str] = []
        i = 0
        while i < len(s):
            ch = s[i]
            if ch.isdigit():
                t = "DIGIT"
            elif ch.isalpha():
                t = "RU" if is_cyrillic_char(ch) else "EN"
            else:
                t = "OTHER"
            j = i + 1
            while j < len(s):
                ch2 = s[j]
                if t == "DIGIT" and ch2.isdigit():
                    j += 1
                elif t == "EN" and ch2.isalpha() and not is_cyrillic_char(ch2):
                    j += 1
                elif t == "RU" and ch2.isalpha() and is_cyrillic_char(ch2):
                    j += 1
                elif t == "OTHER" and (not ch2.isdigit()) and (not ch2.isalpha()):
                    j += 1
                else:
                    break
            chunk = s[i:j]
            if t == "DIGIT":
                parts.append(self.transform_token(chunk, mode, key, tweak, ALPH_DIGITS))
            elif t == "EN":
                parts.append(self.transform_token(chunk, mode, key, tweak, ALPH_EN))
            elif t == "RU":
                parts.append(self.transform_token(chunk, mode, key, tweak, ALPH_RU))
            else:
                parts.append(chunk)
            i = j
        return "".join(parts)


def build_engine(engine_name: str, short_policy: str, rounds: int) -> Engine:
    if engine_name == CloudproofFF1Engine.name:
        return CloudproofFF1Engine(short_policy=short_policy)
    if engine_name == CloudproofFF1AlnumEngine.name:
        return CloudproofFF1AlnumEngine(short_policy=short_policy)
    if engine_name == DemoEngine.name:
        return DemoEngine(rounds=rounds, short_policy=short_policy)
    raise ValueError(f"Unknown engine: {engine_name!r}")


# ----------------------------
# Manifest schema + HMAC
# ----------------------------
@dataclass
class ManifestV2:
    version: int
    engine: str
    engine_params: Dict[str, Any]
    rounds: int
    kdf: str
    file_salt_b64: str
    tweak_mode: str                  # row_index | row_id_column
    row_id_columns: Dict[str, str]   # sheet_norm -> header_norm (only for row_id_column)
    columns: Dict[str, List[str]]    # sheet_norm -> [header_norm, ...]
    header_row: int
    short_policy: str
    encrypted_file_sha256: str
    secret_ref: Dict[str, Any]
    manifest_hmac_b64: Optional[str] = None  # filled after creation

    def to_unsigned_dict(self) -> Dict[str, Any]:
        return {
            "version": self.version,
            "engine": self.engine,
            "engine_params": self.engine_params,
            "rounds": self.rounds,
            "kdf": self.kdf,
            "file_salt_b64": self.file_salt_b64,
            "tweak_mode": self.tweak_mode,
            "row_id_columns": self.row_id_columns,
            "columns": self.columns,
            "header_row": self.header_row,
            "short_policy": self.short_policy,
            "encrypted_file_sha256": self.encrypted_file_sha256,
            "secret_ref": self.secret_ref,
        }

    def to_dict(self) -> Dict[str, Any]:
        d = self.to_unsigned_dict()
        d["manifest_hmac_b64"] = self.manifest_hmac_b64
        return d

    @staticmethod
    def from_dict(d: Dict[str, Any]) -> "ManifestV2":
        return ManifestV2(
            version=int(d.get("version", 2)),
            engine=d.get("engine", DemoEngine.name),
            engine_params=d.get("engine_params", {}),
            rounds=int(d.get("rounds", 10)),
            kdf=d.get("kdf", "HKDF-like(HMAC-SHA256)"),
            file_salt_b64=d["file_salt_b64"],
            tweak_mode=d["tweak_mode"],
            row_id_columns=d.get("row_id_columns", {}),
            columns=d["columns"],
            header_row=int(d.get("header_row", 1)),
            short_policy=d.get("short_policy", "leave"),
            encrypted_file_sha256=d["encrypted_file_sha256"],
            secret_ref=d.get("secret_ref", {}),
            manifest_hmac_b64=d.get("manifest_hmac_b64"),
        )

def derive_column_key(file_secret: bytes, file_salt: bytes, sheet_norm: str, header_norm: str) -> bytes:
    return hkdf_like(file_secret, salt=file_salt, info=f"key|{sheet_norm}|{header_norm}".encode("utf-8"), length=32)

def derive_tweak(file_secret: bytes, file_salt: bytes, sheet_norm: str, header_norm: str, row_tag: str) -> bytes:
    info = f"tweak|{sheet_norm}|{header_norm}|{row_tag}".encode("utf-8")
    return hkdf_like(file_secret, salt=file_salt, info=info, length=16)

def derive_manifest_hmac_key(file_secret: bytes, file_salt: bytes) -> bytes:
    return hkdf_like(file_secret, salt=file_salt, info=b"manifest-hmac-key", length=32)

def sign_manifest(unsigned: Dict[str, Any], file_secret: bytes) -> Tuple[str, str]:
    file_salt = b64d(unsigned["file_salt_b64"])
    mac_key = derive_manifest_hmac_key(file_secret, file_salt)
    msg = canonical_json_bytes(unsigned)
    sig = hmac.new(mac_key, msg, hashlib.sha256).digest()
    return b64e(sig), hashlib.sha256(msg).hexdigest()

def verify_manifest(mf: ManifestV2, file_secret: bytes) -> None:
    if not mf.manifest_hmac_b64:
        raise ValueError("Manifest has no manifest_hmac_b64. Refusing to proceed.")
    unsigned = mf.to_unsigned_dict()
    expected_sig_b64, _ = sign_manifest(unsigned, file_secret)
    if not hmac.compare_digest(expected_sig_b64, mf.manifest_hmac_b64):
        raise ValueError("Manifest HMAC verification failed (manifest may be modified or wrong secret provided).")


# ----------------------------
# XLSX transformation
# ----------------------------
def resolve_sheet_name(wb, sheet_input: str) -> Optional[str]:
    target = norm_text(sheet_input)
    for name in wb.sheetnames:
        if norm_text(name) == target:
            return name
    return None

def header_map(ws, header_row: int) -> Dict[str, int]:
    cells = list(ws[header_row])
    m: Dict[str, int] = {}
    for idx, cell in enumerate(cells, start=1):
        nh = norm_text(cell.value)
        if nh and nh not in m:
            m[nh] = idx
    return m

def transform_workbook(
    in_xlsx: str,
    out_xlsx: str,
    engine: Engine,
    file_secret: bytes,
    file_salt: bytes,
    columns_by_sheet: Dict[str, List[str]],
    tweak_mode: str,
    row_id_columns: Dict[str, str],
    header_row: int,
    mode: str,
    force_string: bool = True,
) -> None:
    wb = load_workbook(in_xlsx)
    for sheet_norm, headers_norm in columns_by_sheet.items():
        actual_sheet = resolve_sheet_name(wb, sheet_norm)
        if actual_sheet is None:
            raise ValueError(f"Sheet not found: {sheet_norm!r}")
        ws = wb[actual_sheet]
        hmap = header_map(ws, header_row=header_row)

        # Resolve row-id column (if needed)
        id_col_idx: Optional[int] = None
        id_header_norm: Optional[str] = None
        if tweak_mode == "row_id_column":
            id_header_norm = row_id_columns.get(sheet_norm)
            if not id_header_norm:
                raise ValueError(f"tweak_mode=row_id_column but no row_id_columns entry for sheet {sheet_norm!r}")
            if id_header_norm not in hmap:
                raise ValueError(f"Row-id column {id_header_norm!r} not found in sheet {sheet_norm!r}")
            id_col_idx = hmap[id_header_norm]
            if id_header_norm in set(headers_norm):
                raise ValueError(
                    f"Row-id column {id_header_norm!r} is also in encrypted columns for sheet {sheet_norm!r}. "
                    f"Choose an ID column that is NOT encrypted."
                )

        for header_norm in headers_norm:
            if header_norm not in hmap:
                raise ValueError(f"Column header not found on sheet {sheet_norm!r}: {header_norm!r}")
            col_idx = hmap[header_norm]
            col_key = derive_column_key(file_secret, file_salt, sheet_norm, header_norm)

            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is None:
                    continue

                if tweak_mode == "row_index":
                    row_tag = f"row:{r}"
                else:
                    assert id_col_idx is not None
                    id_val = ws.cell(row=r, column=id_col_idx).value
                    if id_val is None or norm_text(id_val) == "":
                        raise ValueError(f"Empty row-id value at sheet={sheet_norm!r}, row={r}, id_col={id_header_norm!r}")
                    id_text = norm_text(id_val).encode("utf-8")
                    id_hash = hashlib.sha256(id_text).hexdigest()
                    row_tag = f"id:{id_hash}"

                tweak = derive_tweak(file_secret, file_salt, sheet_norm, header_norm, row_tag)
                new_val = engine.transform_cell_value(cell.value, mode=mode, key=col_key, tweak=tweak)
                cell.value = new_val

                # Optional: force everything to string to avoid numeric precision / leading-zero surprises
                if force_string:
                    cell.data_type = "s"

    wb.save(out_xlsx)


# ----------------------------
# Config schema
# ----------------------------
def load_config(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    columns_raw = cfg.get("columns")
    if not isinstance(columns_raw, dict) or not columns_raw:
        raise ValueError("Config must contain non-empty object 'columns' (sheet -> [headers...]).")

    columns_by_sheet: Dict[str, List[str]] = {}
    for sheet, headers in columns_raw.items():
        sheet_n = norm_text(sheet)
        if not isinstance(headers, list) or not headers:
            raise ValueError(f"Config columns for sheet {sheet!r} must be a non-empty list.")
        headers_n = [norm_text(h) for h in headers]
        columns_by_sheet[sheet_n] = headers_n

    tweak_mode = cfg.get("tweak_mode", "row_id_column")
    if tweak_mode not in ("row_id_column", "row_index"):
        raise ValueError("tweak_mode must be 'row_id_column' or 'row_index'.")

    row_id_columns: Dict[str, str] = {}
    if tweak_mode == "row_id_column":
        rid = cfg.get("row_id_columns")
        if not isinstance(rid, dict) or not rid:
            raise ValueError("With tweak_mode=row_id_column you must set 'row_id_columns' (sheet -> id_header).")
        for sheet, header in rid.items():
            row_id_columns[norm_text(sheet)] = norm_text(header)

    header_row = int(cfg.get("header_row", 1))
    short_policy = cfg.get("short_policy", "leave")
    if short_policy not in ("leave", "error"):
        raise ValueError("short_policy must be 'leave' or 'error'.")

    engine = cfg.get("engine", CloudproofFF1Engine.name)
    rounds = int(cfg.get("rounds", 10))  # only used for demo engine
    force_string = bool(cfg.get("force_string", True))

    return {
        "columns_by_sheet": columns_by_sheet,
        "tweak_mode": tweak_mode,
        "row_id_columns": row_id_columns,
        "header_row": header_row,
        "short_policy": short_policy,
        "engine": engine,
        "rounds": rounds,
        "force_string": force_string,
        "engine_params": cfg.get("engine_params", {}),
    }


# ----------------------------
# CLI
# ----------------------------
def cmd_encrypt(args: argparse.Namespace) -> int:
    cfg = load_config(args.config)

    file_secret = b64d(args.secret_b64) if args.secret_b64 else os.urandom(32)
    file_salt = os.urandom(16)

    engine = build_engine(cfg["engine"], short_policy=cfg["short_policy"], rounds=cfg["rounds"])

    transform_workbook(
        in_xlsx=args.input,
        out_xlsx=args.output,
        engine=engine,
        file_secret=file_secret,
        file_salt=file_salt,
        columns_by_sheet=cfg["columns_by_sheet"],
        tweak_mode=cfg["tweak_mode"],
        row_id_columns=cfg["row_id_columns"],
        header_row=cfg["header_row"],
        mode="encrypt",
        force_string=cfg["force_string"],
    )

    enc_hash = sha256_file(args.output)

    mf = ManifestV2(
        version=2,
        engine=engine.name,
        engine_params=cfg.get("engine_params", {}),
        rounds=getattr(engine, "rounds", 18) if engine.name == DemoEngine.name else 18,
        kdf="HKDF-like(HMAC-SHA256)",
        file_salt_b64=b64e(file_salt),
        tweak_mode=cfg["tweak_mode"],
        row_id_columns=cfg["row_id_columns"],
        columns=cfg["columns_by_sheet"],
        header_row=cfg["header_row"],
        short_policy=cfg["short_policy"],
        encrypted_file_sha256=enc_hash,
        secret_ref={"type": "one_time_channel", "id": args.secret_ref or ""},
        manifest_hmac_b64=None,
    )

    unsigned = mf.to_unsigned_dict()
    sig_b64, _canon_sha = sign_manifest(unsigned, file_secret)
    mf.manifest_hmac_b64 = sig_b64

    with open(args.manifest, "w", encoding="utf-8") as f:
        json.dump(mf.to_dict(), f, ensure_ascii=False, indent=2)

    if args.secret_out:
        with open(args.secret_out, "w", encoding="utf-8") as f:
            f.write(b64e(file_secret))
    else:
        print(b64e(file_secret))

    print(f"[OK] Encrypted: {args.output}")
    print(f"[OK] Manifest : {args.manifest}")
    print(f"[OK] Encrypted SHA256: {enc_hash}")
    return 0

def cmd_decrypt(args: argparse.Namespace) -> int:
    with open(args.manifest, "r", encoding="utf-8") as f:
        mf = ManifestV2.from_dict(json.load(f))

    file_secret = b64d(args.secret_b64) if args.secret_b64 else b64d(open(args.secret_file, "r", encoding="utf-8").read().strip())
    file_salt = b64d(mf.file_salt_b64)

    enc_hash = sha256_file(args.input)
    if enc_hash != mf.encrypted_file_sha256:
        raise ValueError(f"Encrypted file SHA256 mismatch. Expected {mf.encrypted_file_sha256}, got {enc_hash}.")

    verify_manifest(mf, file_secret)

    engine = build_engine(mf.engine, short_policy=mf.short_policy, rounds=mf.rounds)

    transform_workbook(
        in_xlsx=args.input,
        out_xlsx=args.output,
        engine=engine,
        file_secret=file_secret,
        file_salt=file_salt,
        columns_by_sheet=mf.columns,
        tweak_mode=mf.tweak_mode,
        row_id_columns=mf.row_id_columns,
        header_row=mf.header_row,
        mode="decrypt",
        force_string=True,  # safest on decrypt too
    )

    print(f"[OK] Decrypted: {args.output}")
    return 0

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="XLSX column format-preserving encryption with manifest.")
    sub = p.add_subparsers(dest="cmd", required=True)

    pe = sub.add_parser("encrypt", help="Encrypt selected columns in XLSX")
    pe.add_argument("--in", dest="input", required=True, help="Input XLSX")
    pe.add_argument("--out", dest="output", required=True, help="Output XLSX (encrypted)")
    pe.add_argument("--manifest", required=True, help="Output manifest.json")
    pe.add_argument("--config", required=True, help="Config JSON (sheets/columns/tweak mode)")
    pe.add_argument("--secret-out", help="Write base64 file_secret to this file (optional). If omitted, prints to stdout.")
    pe.add_argument("--secret-b64", help="Provide your own base64 file_secret (optional). If omitted, generated randomly.")
    pe.add_argument("--secret-ref", help="Optional ID for your secret channel (stored in manifest).")

    pd = sub.add_parser("decrypt", help="Decrypt XLSX using manifest + file_secret")
    pd.add_argument("--in", dest="input", required=True, help="Input XLSX (encrypted)")
    pd.add_argument("--out", dest="output", required=True, help="Output XLSX (decrypted)")
    pd.add_argument("--manifest", required=True, help="manifest.json from encryption step")
    g = pd.add_mutually_exclusive_group(required=True)
    g.add_argument("--secret-b64", help="base64 file_secret")
    g.add_argument("--secret-file", help="Path to file containing base64 file_secret")

    return p

def main() -> int:
    args = build_parser().parse_args()
    if args.cmd == "encrypt":
        return cmd_encrypt(args)
    if args.cmd == "decrypt":
        return cmd_decrypt(args)
    raise RuntimeError("Unknown command")

if __name__ == "__main__":
    raise SystemExit(main())
