#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import io
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import column_index_from_string, get_column_letter


# ----------------------------
# Config models
# ----------------------------

@dataclass(frozen=True)
class NormalizeConfig:
    strip: bool = True
    casefold: bool = True


@dataclass(frozen=True)
class TokenizerConfig:
    header_row: int = 1
    token_start: int = 1
    normalize: NormalizeConfig = NormalizeConfig()
    targets: List[Dict[str, Any]] = None  # {"sheet": str, "columns": [str, ...]}


# ----------------------------
# Mapping models
# ----------------------------

@dataclass
class ColumnMap:
    sheet: str
    column_letter: str                 # resolved column letter (e.g., "C")
    column_ref: str                    # original ref from config (e.g., "Status" or "C")
    token_start: int
    values: List[str]                  # ordered by alphabetical enumeration


@dataclass
class MappingFile:
    version: int
    workbook_hint: Dict[str, Any]
    normalize: Dict[str, Any]
    maps: List[ColumnMap]


# ----------------------------
# Helpers
# ----------------------------

def die(msg: str, code: int = 2) -> None:
    print(f"ERROR: {msg}", file=sys.stderr)
    raise SystemExit(code)


def load_config(path: Path) -> TokenizerConfig:
    data = json.loads(path.read_text(encoding="utf-8"))
    if "targets" not in data or not isinstance(data["targets"], list):
        die("config.json: поле 'targets' обязательно и должно быть списком")

    norm = data.get("normalize", {}) or {}
    normalize = NormalizeConfig(
        strip=bool(norm.get("strip", True)),
        casefold=bool(norm.get("casefold", True)),
    )

    header_row = int(data.get("header_row", 1))
    token_start = int(data.get("token_start", 1))
    targets = data["targets"]

    for t in targets:
        if "sheet" not in t or "columns" not in t:
            die("config.json: каждый элемент targets должен иметь 'sheet' и 'columns'")
        if not isinstance(t["columns"], list) or not t["columns"]:
            die("config.json: 'columns' должен быть непустым списком")

    return TokenizerConfig(
        header_row=header_row,
        token_start=token_start,
        normalize=normalize,
        targets=targets,
    )


def normalize_value(v: Any, norm: NormalizeConfig) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v
    else:
        s = str(v)

    if norm.strip:
        s = s.strip()

    if s == "":
        return None
    return s


def sort_key_for_enum(s: str, norm: NormalizeConfig) -> str:
    return s.casefold() if norm.casefold else s


def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def load_workbook_from_bytes(b: bytes) -> Workbook:
    return load_workbook(io.BytesIO(b))


def resolve_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    if sheet_name not in wb.sheetnames:
        die(f"Лист '{sheet_name}' не найден. Доступные: {wb.sheetnames}")
    return wb[sheet_name]


def is_column_letter_ref(ref: str) -> bool:
    r = ref.strip()
    return r.isalpha() and 1 <= len(r) <= 3


def resolve_column_letter(ws: Worksheet, col_ref: str, header_row: int, norm: NormalizeConfig) -> str:
    ref = col_ref.strip()

    # If explicitly a column letter like "A", "BC"
    if is_column_letter_ref(ref) and ref.upper() == ref:
        try:
            _ = column_index_from_string(ref)
            return ref
        except Exception:
            pass

    # Otherwise resolve by header name
    target = normalize_value(ref, norm)
    if target is None:
        die(f"Пустой col_ref в конфиге для листа '{ws.title}'")

    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        val = normalize_value(cell.value, norm)
        if val is None:
            continue
        if (val.casefold() if norm.casefold else val) == (target.casefold() if norm.casefold else target):
            return get_column_letter(c)

    die(f"Не найден столбец по заголовку '{col_ref}' на листе '{ws.title}' (строка заголовка: {header_row})")
    return "A"  # unreachable


def iter_column_cells(ws: Worksheet, column_letter: str, start_row: int) -> List[Tuple[int, Any]]:
    col_idx = column_index_from_string(column_letter)
    out: List[Tuple[int, Any]] = []
    for r in range(start_row, ws.max_row + 1):
        out.append((r, ws.cell(row=r, column=col_idx).value))
    return out


def collect_enum_from_column(ws: Worksheet, column_letter: str, data_start_row: int, norm: NormalizeConfig) -> List[str]:
    seen_norm: Dict[str, str] = {}
    for r, raw in iter_column_cells(ws, column_letter, data_start_row):
        v = normalize_value(raw, norm)
        if v is None:
            continue
        k = v.casefold() if norm.casefold else v
        if k not in seen_norm:
            seen_norm[k] = v

    values = list(seen_norm.values())
    values.sort(key=lambda s: sort_key_for_enum(s, norm))
    return values


# ----------------------------
# Core: encode / decode / verify
# ----------------------------

def encode_workbook(wb: Workbook, cfg: TokenizerConfig) -> Tuple[Workbook, MappingFile]:
    maps: List[ColumnMap] = []

    for t in cfg.targets:
        sheet_name = t["sheet"]
        ws = resolve_sheet(wb, sheet_name)
        data_start_row = cfg.header_row + 1

        for col_ref in t["columns"]:
            col_letter = resolve_column_letter(ws, str(col_ref), cfg.header_row, cfg.normalize)
            enum_values = collect_enum_from_column(ws, col_letter, data_start_row, cfg.normalize)

            value_to_token: Dict[str, int] = {}
            for i, v in enumerate(enum_values):
                key = v.casefold() if cfg.normalize.casefold else v
                value_to_token[key] = cfg.token_start + i

            col_idx = column_index_from_string(col_letter)
            for r in range(data_start_row, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                v = normalize_value(cell.value, cfg.normalize)
                if v is None:
                    continue
                k = v.casefold() if cfg.normalize.casefold else v
                if k not in value_to_token:
                    die(f"Значение '{v}' (лист {sheet_name}, {col_letter}{r}) отсутствует в enum")
                cell.value = value_to_token[k]

            maps.append(ColumnMap(
                sheet=sheet_name,
                column_letter=col_letter,
                column_ref=str(col_ref),
                token_start=cfg.token_start,
                values=enum_values,
            ))

    mapping = MappingFile(
        version=1,
        workbook_hint={},
        normalize={"strip": cfg.normalize.strip, "casefold": cfg.normalize.casefold},
        maps=maps,
    )
    return wb, mapping


def decode_workbook(wb: Workbook, mapping: MappingFile, header_row: int = 1) -> Workbook:
    for m in mapping.maps:
        ws = resolve_sheet(wb, m.sheet)
        col_idx = column_index_from_string(m.column_letter)
        data_start_row = header_row + 1

        token_to_value: Dict[int, str] = {m.token_start + i: v for i, v in enumerate(m.values)}

        for r in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is None or cell.value == "":
                continue

            token: Optional[int] = None
            if isinstance(cell.value, int):
                token = cell.value
            elif isinstance(cell.value, float) and cell.value.is_integer():
                token = int(cell.value)
            elif isinstance(cell.value, str):
                s = cell.value.strip()
                if s.isdigit():
                    token = int(s)

            if token is None:
                continue

            if token not in token_to_value:
                die(f"Неизвестный токен {token} (лист {m.sheet}, {m.column_letter}{r})")
            cell.value = token_to_value[token]

    return wb


def mapping_to_json_dict(m: MappingFile, source_filename: str) -> Dict[str, Any]:
    return {
        "version": m.version,
        "workbook_hint": {"source_filename": source_filename},
        "normalize": m.normalize,
        "maps": [
            {
                "sheet": cm.sheet,
                "column_letter": cm.column_letter,
                "column_ref": cm.column_ref,
                "token_start": cm.token_start,
                "values": cm.values,
            }
            for cm in m.maps
        ]
    }


def mapping_from_json_dict(d: Dict[str, Any]) -> MappingFile:
    if int(d.get("version", 0)) != 1:
        die("mapping.json: неподдерживаемая версия")
    maps = []
    for item in d.get("maps", []):
        maps.append(ColumnMap(
            sheet=item["sheet"],
            column_letter=item["column_letter"],
            column_ref=item.get("column_ref", item["column_letter"]),
            token_start=int(item.get("token_start", 1)),
            values=list(item["values"]),
        ))
    normalize = d.get("normalize", {"strip": True, "casefold": True})
    return MappingFile(
        version=1,
        workbook_hint=d.get("workbook_hint", {}),
        normalize=normalize,
        maps=maps,
    )


def verify_roundtrip(input_path: Path, cfg: TokenizerConfig) -> bool:
    wb_orig = load_workbook(input_path)
    orig_bytes = workbook_to_bytes(wb_orig)

    wb_enc = load_workbook_from_bytes(orig_bytes)
    wb_enc, mapping = encode_workbook(wb_enc, cfg)

    enc_bytes = workbook_to_bytes(wb_enc)
    wb_dec = load_workbook_from_bytes(enc_bytes)
    wb_dec = decode_workbook(wb_dec, mapping, header_row=cfg.header_row)

    ok = True
    for t in cfg.targets:
        sheet = t["sheet"]
        ws_o = resolve_sheet(wb_orig, sheet)
        ws_d = resolve_sheet(wb_dec, sheet)
        data_start = cfg.header_row + 1

        for col_ref in t["columns"]:
            col_letter = resolve_column_letter(ws_o, str(col_ref), cfg.header_row, cfg.normalize)
            col_idx = column_index_from_string(col_letter)

            max_row = max(ws_o.max_row, ws_d.max_row)
            for r in range(data_start, max_row + 1):
                vo = normalize_value(ws_o.cell(row=r, column=col_idx).value, cfg.normalize)
                vd = normalize_value(ws_d.cell(row=r, column=col_idx).value, cfg.normalize)
                if vo != vd:
                    ok = False
                    coord = f"{col_letter}{r}"
                    print(f"[MISMATCH] sheet={sheet} cell={coord} original={ws_o.cell(r, col_idx).value!r} decoded={ws_d.cell(r, col_idx).value!r}")

    return ok


# ----------------------------
# CLI
# ----------------------------

def cmd_encode(args: argparse.Namespace) -> None:
    cfg = load_config(Path(args.config))
    wb = load_workbook(args.input)
    wb, mapping = encode_workbook(wb, cfg)
    wb.save(args.output)

    mapping_dict = mapping_to_json_dict(mapping, source_filename=str(args.input))
    Path(args.mapping).write_text(json.dumps(mapping_dict, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK: encoded saved to {args.output}")
    print(f"OK: mapping saved to {args.mapping}")


def cmd_decode(args: argparse.Namespace) -> None:
    cfg = load_config(Path(args.config))
    wb = load_workbook(args.input)

    mapping_dict = json.loads(Path(args.mapping).read_text(encoding="utf-8"))
    mapping = mapping_from_json_dict(mapping_dict)

    wb = decode_workbook(wb, mapping, header_row=cfg.header_row)
    wb.save(args.output)

    print(f"OK: decoded saved to {args.output}")


def cmd_verify(args: argparse.Namespace) -> None:
    cfg = load_config(Path(args.config))
    ok = verify_roundtrip(Path(args.input), cfg)
    if ok:
        print("OK: verify passed (encode -> decode restores original values in targeted columns)")
        raise SystemExit(0)
    else:
        print("FAIL: verify found mismatches")
        raise SystemExit(1)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="xlsxtok",
        description="Tokenize/Detokenize XLSX columns by alphabetical enumeration of unique values in column."
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    e = sub.add_parser("encode", help="Tokenize selected columns (values -> ordinal tokens)")
    e.add_argument("-i", "--input", required=True, help="Input .xlsx")
    e.add_argument("-c", "--config", required=True, help="config.json")
    e.add_argument("-o", "--output", required=True, help="Output tokenized .xlsx")
    e.add_argument("-m", "--mapping", required=True, help="Output mapping.json")
    e.set_defaults(func=cmd_encode)

    d = sub.add_parser("decode", help="Detokenize selected columns (tokens -> original values)")
    d.add_argument("-i", "--input", required=True, help="Input tokenized .xlsx")
    d.add_argument("-c", "--config", required=True, help="config.json (for header_row)")
    d.add_argument("-m", "--mapping", required=True, help="mapping.json from encode")
    d.add_argument("-o", "--output", required=True, help="Output decoded .xlsx")
    d.set_defaults(func=cmd_decode)

    v = sub.add_parser("verify", help="Check that encode->decode restores original values")
    v.add_argument("-i", "--input", required=True, help="Input original .xlsx")
    v.add_argument("-c", "--config", required=True, help="config.json")
    v.set_defaults(func=cmd_verify)

    return p


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
