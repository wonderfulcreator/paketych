#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ETA regression core (stage-wise) with a callable API.

What it does
- Loads an Excel table of shipments.
- Builds 3 staged regression models to predict remaining time (days) to final port:
    Stage1: after departure from origin port
    Stage2: after arrival to transshipment port
    Stage3: after departure from transshipment port
- Evaluates on a holdout year (default 2025).
- Optionally retrains on all facts (for best forward predictions).
- Appends predictions back to the Excel + applies clean formatting.

Public API
    run_pipeline(input_path, output_path, test_year=2025, no_retrain_all=False, logger=None)

"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Callable

import math
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder
from sklearn.linear_model import Ridge
from sklearn.metrics import mean_absolute_error, mean_squared_error


# -----------------------------
# Columns & aliases
# -----------------------------

DATE_COLS = [
    'Дата готовности груза',
    'Перевозка.Дата выпуска заявления',
    'Перевозка.Дата выпуска ДТ',
    'Дата выхода из порта вывоза план',
    'Дата выхода из порта вывоза расчет',
    'Дата выхода из порта вывоза слежение',
    'Дата выхода из порта вывоза факт',
    'Дата прихода в порт перегрузки план',
    'Дата прихода в порт перегрузки расчет',
    'Дата прихода в порт перегрузки слежение',
    'Дата прихода в порт перегрузки факт',
    'Дата выхода из порта перегрузки план',
    'Дата выхода из порта перегрузки расчет',
    'Дата выхода из порта перегрузки слежение',
    'Дата выхода из порта перегрузки факт',
    'Дата прихода в конечный порт план',
    'Дата прихода в конечный порт расчет',
    'Дата прихода в конечный порт слежение',
    'Дата прихода в конечный порт факт',
]

TARGET_FINAL_FACT = 'Дата прихода в конечный порт факт'

# Short, user-friendly column names for output (and accepted as input aliases)
COL_ALIASES: Dict[str, str] = {
    # Route fields
    'Порт вывоза': 'Порт_вывоза',
    'Порт перегрузки': 'Порт_перегрузки',
    'Конечный порт': 'Порт_конечный',
    'Заказ поставщику.Условия поставки': 'Условия',
    'Перевозка.Морская линия': 'Линия',
    'Перевозка.Судно': 'Судно',
    'Перевозка.Экспедитор': 'Экспедитор',

    # Facts
    'Дата выхода из порта вывоза факт': 'Факт_выход_вывоз',
    'Дата прихода в порт перегрузки факт': 'Факт_приход_перегруз',
    'Дата выхода из порта перегрузки факт': 'Факт_выход_перегруз',
    'Дата прихода в конечный порт факт': 'Факт_приход_конечный',
    'Дата прихода в конечный порт расчет': 'План_ETA_конечный',

    # Predictions
    'ETA_прогноз_stage1': 'ETA_модель_S1',
    'ETA_прогноз_stage2': 'ETA_модель_S2',
    'ETA_прогноз_stage3': 'ETA_модель_S3',
    'ETA_прогноз_auto': 'ETA_модель',
    'ETA_прогноз_auto_stage': 'Стадия_модели',
    'ETA_прогноз_auto_remaining_days': 'Остаток_дней',
}

REVERSE_ALIASES: Dict[str, str] = {v: k for k, v in COL_ALIASES.items()}


def _normalize_input_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Accept both canonical and short column names as input."""
    rename_map = {}
    for short, canonical in REVERSE_ALIASES.items():
        if canonical not in df.columns and short in df.columns:
            rename_map[short] = canonical
    return df.rename(columns=rename_map) if rename_map else df


def _apply_output_aliases(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to short names for the exported (user-facing) Excel."""
    rename_map = {c: COL_ALIASES[c] for c in df.columns if c in COL_ALIASES}
    return df.rename(columns=rename_map)


CAT_BASE = [
    'Поставщик',
    'Порт вывоза',
    'Порт перегрузки',
    'Конечный порт',
    'Заказ поставщику.Условия поставки',
    'Перевозка.Морская линия',
    'Перевозка.Судно',
    'Перевозка.Экспедитор',
]


def _to_datetime(df: pd.DataFrame) -> pd.DataFrame:
    for c in DATE_COLS:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], dayfirst=True, errors='coerce')
    return df


def _save_excel_pretty(df: pd.DataFrame, path: str) -> None:
    """Save df to Excel with compact date display + clean formatting.

    Important: we only set date formats on *true* datetime columns.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        wb = writer.book
        ws = wb["Data"]

        ws.freeze_panes = "A2"
        ws.sheet_view.zoomScale = 110

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.row_dimensions[1].height = 24

        ws.auto_filter.ref = ws.dimensions

        # Column widths
        for col_idx, col_name in enumerate(df.columns, start=1):
            s = df[col_name]
            sample = s.head(200).astype(str).replace({"nan": "", "NaT": ""})
            max_len = max([len(str(col_name))] + [len(x) for x in sample.tolist() if x is not None])
            width = min(40, max(10, max_len + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Formats
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_series = df[col_name]
            col_str = str(col_name)

            if pd.api.types.is_datetime64_any_dtype(col_series):
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = "DD.MM.YYYY"

            # Remaining days must stay numeric
            elif ("Остаток" in col_str) or (col_str.endswith("_days")) or ("remaining" in col_str.lower()):
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = "0"

        # Table style
        try:
            table = Table(displayName="ETA_Table", ref=ws.dimensions)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        except Exception:
            pass


def _route_signature(df: pd.DataFrame) -> pd.Series:
    return (
        df['Порт вывоза'].astype(str)
        + '>'
        + df['Порт перегрузки'].astype(str)
        + '>'
        + df['Конечный порт'].astype(str)
    )


def _safe_days(delta: pd.Series) -> pd.Series:
    return delta.dt.total_seconds() / 86400.0


def _fill_cat_as_str(X: pd.DataFrame, cat_cols: List[str]) -> pd.DataFrame:
    for c in cat_cols:
        if c in X.columns:
            X[c] = X[c].where(X[c].notna(), 'MISSING').astype(str)
    return X


@dataclass
class StageSpec:
    name: str
    ref_col: str
    requires_transshipment: bool
    numeric_cols: List[str]


STAGES: Dict[str, StageSpec] = {
    'stage1': StageSpec(
        name='stage1',
        ref_col='Дата выхода из порта вывоза факт',
        requires_transshipment=False,
        numeric_cols=[
            'route_mean_days',
            'ref_month',
            'ref_dow',
            'planned_remaining_calc',
            'planned_remaining_plan',
        ],
    ),
    'stage2': StageSpec(
        name='stage2',
        ref_col='Дата прихода в порт перегрузки факт',
        requires_transshipment=True,
        numeric_cols=[
            'route_mean_days',
            'elapsed_origin_to_ref',
            'ref_month',
            'ref_dow',
            'planned_remaining_calc',
            'planned_remaining_plan',
            'planned_wait_transship_calc',
            'planned_wait_transship_plan',
        ],
    ),
    'stage3': StageSpec(
        name='stage3',
        ref_col='Дата выхода из порта перегрузки факт',
        requires_transshipment=True,
        numeric_cols=[
            'route_mean_days',
            'elapsed_origin_to_ref',
            'dwell_transship',
            'ref_month',
            'ref_dow',
            'planned_remaining_calc',
            'planned_remaining_plan',
        ],
    ),
}


class ETAModelStage:
    def __init__(self, spec: StageSpec):
        self.spec = spec
        self.cat_cols = CAT_BASE
        self.num_cols = spec.numeric_cols
        self.route_mean_map: Dict[str, float] = {}
        self.route_global_mean: float = float('nan')
        self.num_medians: Dict[str, float] = {}
        self.model: Optional[Pipeline] = None

    def _build_features(self, df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        out['route_signature'] = _route_signature(out)
        ref = out[self.spec.ref_col]

        out['ref_month'] = ref.dt.month.astype('float')
        out['ref_dow'] = ref.dt.dayofweek.astype('float')

        out['planned_remaining_calc'] = _safe_days(out['Дата прихода в конечный порт расчет'] - ref)
        out['planned_remaining_plan'] = _safe_days(out['Дата прихода в конечный порт план'] - ref)

        if self.spec.name != 'stage1':
            out['elapsed_origin_to_ref'] = _safe_days(ref - out['Дата выхода из порта вывоза факт'])

        if self.spec.name == 'stage2':
            out['planned_wait_transship_calc'] = _safe_days(out['Дата выхода из порта перегрузки расчет'] - ref)
            out['planned_wait_transship_plan'] = _safe_days(out['Дата выхода из порта перегрузки план'] - ref)

        if self.spec.name == 'stage3':
            out['dwell_transship'] = _safe_days(ref - out['Дата прихода в порт перегрузки факт'])

        out['route_mean_days'] = out['route_signature'].map(self.route_mean_map).astype('float')
        out['route_mean_days'] = out['route_mean_days'].fillna(self.route_global_mean)

        return out

    def _mask_applicable(self, df: pd.DataFrame) -> pd.Series:
        m = df[self.spec.ref_col].notna()
        if self.spec.requires_transshipment:
            m = m & df['Порт перегрузки'].notna()
        return m

    def _prepare_X(self, feat: pd.DataFrame) -> pd.DataFrame:
        X = feat[self.cat_cols + self.num_cols].copy()
        X = _fill_cat_as_str(X, self.cat_cols)

        for c, med in self.num_medians.items():
            if c in X.columns:
                X[c] = X[c].astype('float').fillna(med)

        for c in self.num_cols:
            if c in X.columns and X[c].isna().any():
                med = float(np.nanmedian(X[c].values))
                if np.isnan(med):
                    med = 0.0
                X[c] = X[c].fillna(med)
        return X

    def fit(self, df: pd.DataFrame, holdout_year: Optional[int] = None) -> Tuple[Optional[Dict[str, float]], int]:
        m = self._mask_applicable(df) & df[TARGET_FINAL_FACT].notna()
        if holdout_year is not None:
            m = m & (df[TARGET_FINAL_FACT].dt.year != holdout_year)
        work = df.loc[m].copy()
        if len(work) < 50:
            self.model = None
            return None, 0

        y = _safe_days(work[TARGET_FINAL_FACT] - work[self.spec.ref_col]).astype(float)

        work['route_signature'] = _route_signature(work)
        work['_y'] = y.values
        route_mean = work.groupby('route_signature')['_y'].mean()
        self.route_mean_map = route_mean.to_dict()
        self.route_global_mean = float(np.nanmean(y.values))

        feat = self._build_features(work)
        X = feat[self.cat_cols + self.num_cols].copy()
        X = _fill_cat_as_str(X, self.cat_cols)

        self.num_medians = {}
        for c in self.num_cols:
            vals = pd.to_numeric(X[c], errors='coerce')
            med = float(np.nanmedian(vals.values))
            if np.isnan(med):
                med = 0.0
            self.num_medians[c] = med
            X[c] = vals.fillna(med).astype(float)

        preprocessor = ColumnTransformer(
            transformers=[
                ('cat', OneHotEncoder(handle_unknown='ignore'), self.cat_cols),
                ('num', 'passthrough', self.num_cols),
            ]
        )

        self.model = Pipeline(
            steps=[
                ('preprocess', preprocessor),
                ('model', Ridge(alpha=10.0, random_state=42)),
            ]
        )
        self.model.fit(X, y)
        return self.num_medians, len(work)

    def evaluate(self, df: pd.DataFrame, test_year: int) -> Optional[Dict[str, float]]:
        if self.model is None:
            return None

        m = self._mask_applicable(df) & df[TARGET_FINAL_FACT].notna() & (df[TARGET_FINAL_FACT].dt.year == test_year)
        if m.sum() == 0:
            return None

        work = df.loc[m].copy()
        feat = self._build_features(work)
        X = self._prepare_X(feat)
        y_true = _safe_days(work[TARGET_FINAL_FACT] - work[self.spec.ref_col]).astype(float)
        y_pred = self.model.predict(X)

        return {
            'n': int(m.sum()),
            'mae': float(mean_absolute_error(y_true, y_pred)),
            'rmse': float(math.sqrt(mean_squared_error(y_true, y_pred))),
            'median_ae': float(np.median(np.abs(y_true.values - y_pred))),
            'p90_ae': float(np.quantile(np.abs(y_true.values - y_pred), 0.9)),
        }

    def baseline(self, df: pd.DataFrame, test_year: int) -> Optional[Dict[str, float]]:
        ref = df[self.spec.ref_col]
        base = _safe_days(df['Дата прихода в конечный порт расчет'] - ref)
        m = (
            self._mask_applicable(df)
            & df[TARGET_FINAL_FACT].notna()
            & (df[TARGET_FINAL_FACT].dt.year == test_year)
            & base.notna()
        )
        if m.sum() == 0:
            return None
        y_true = _safe_days(df.loc[m, TARGET_FINAL_FACT] - df.loc[m, self.spec.ref_col]).astype(float)
        y_pred = base.loc[m].astype(float)
        return {
            'n': int(m.sum()),
            'mae': float(mean_absolute_error(y_true, y_pred)),
            'rmse': float(math.sqrt(mean_squared_error(y_true, y_pred))),
        }

    def predict_remaining_days(self, df: pd.DataFrame) -> pd.Series:
        if self.model is None:
            return pd.Series([np.nan] * len(df), index=df.index)

        m = self._mask_applicable(df)
        out = pd.Series([np.nan] * len(df), index=df.index, dtype=float)
        if m.sum() == 0:
            return out

        work = df.loc[m].copy()
        feat = self._build_features(work)
        X = self._prepare_X(feat)
        pred = self.model.predict(X)
        out.loc[m] = pred
        return out


def _auto_stage(row: pd.Series) -> str:
    if pd.notna(row.get('Дата выхода из порта перегрузки факт')) and pd.notna(row.get('Порт перегрузки')):
        return 'stage3'
    if pd.notna(row.get('Дата прихода в порт перегрузки факт')) and pd.notna(row.get('Порт перегрузки')):
        return 'stage2'
    return 'stage1'


Logger = Optional[Callable[[str], None]]


def run_pipeline(
    input_path: str,
    output_path: str,
    test_year: int = 2025,
    no_retrain_all: bool = False,
    logger: Logger = None,
) -> Dict[str, Dict[str, float]]:
    """Run the full ETA pipeline.

    Returns a dict with evaluation metrics per stage (baseline + model) for the holdout test_year.
    """

    def log(msg: str) -> None:
        if logger is not None:
            logger(msg)
        else:
            print(msg)

    log(f"Загрузка: {input_path}")
    df = pd.read_excel(input_path)
    df = _normalize_input_columns(df)
    df = _to_datetime(df)

    # Train models excluding test year
    stage_models = {name: ETAModelStage(spec) for name, spec in STAGES.items()}
    for name in ['stage1', 'stage2', 'stage3']:
        _, n = stage_models[name].fit(df, holdout_year=test_year)
        log(f"Обучение {name}: строк в train = {n}")

    # Evaluate
    metrics: Dict[str, Dict[str, float]] = {}
    log(f"\n=== Holdout evaluation on {test_year} (target: {TARGET_FINAL_FACT}) ===")
    for name in ['stage1', 'stage2', 'stage3']:
        sm = stage_models[name]
        base = sm.baseline(df, test_year)
        ev = sm.evaluate(df, test_year)
        metrics[name] = {
            **({f"baseline_{k}": v for k, v in base.items()} if base else {}),
            **({f"model_{k}": v for k, v in ev.items()} if ev else {}),
        }
        if base:
            log(f"[{name}] Baseline(calc): n={base['n']} MAE={base['mae']:.2f}d RMSE={base['rmse']:.2f}d")
        else:
            log(f"[{name}] Baseline(calc): n=0")
        if ev:
            log(
                f"[{name}] Model:        n={ev['n']} MAE={ev['mae']:.2f}d RMSE={ev['rmse']:.2f}d "
                f"MedianAE={ev['median_ae']:.2f}d P90AE={ev['p90_ae']:.2f}d"
            )
        else:
            log(f"[{name}] Model:        n=0")

    # Optionally retrain on all data
    if not no_retrain_all:
        log("\nПереобучение на всех данных (для лучшего прогноза вперёд)...")
        for name in ['stage1', 'stage2', 'stage3']:
            stage_models[name].fit(df, holdout_year=None)

    # Predict remaining days
    pred1 = stage_models['stage1'].predict_remaining_days(df)
    pred2 = stage_models['stage2'].predict_remaining_days(df)
    pred3 = stage_models['stage3'].predict_remaining_days(df)

    eta1 = (df[STAGES['stage1'].ref_col] + pd.to_timedelta(pred1.round(), unit='D')).dt.normalize()
    eta2 = (df[STAGES['stage2'].ref_col] + pd.to_timedelta(pred2.round(), unit='D')).dt.normalize()
    eta3 = (df[STAGES['stage3'].ref_col] + pd.to_timedelta(pred3.round(), unit='D')).dt.normalize()

    auto_stage = df.apply(_auto_stage, axis=1)
    eta_auto = pd.Series([pd.NaT] * len(df), index=df.index)
    rem_auto = pd.Series([np.nan] * len(df), index=df.index, dtype=float)

    mask3 = auto_stage == 'stage3'
    mask2 = auto_stage == 'stage2'
    mask1 = auto_stage == 'stage1'

    eta_auto.loc[mask3] = eta3.loc[mask3]
    rem_auto.loc[mask3] = pred3.loc[mask3]
    eta_auto.loc[mask2] = eta2.loc[mask2]
    rem_auto.loc[mask2] = pred2.loc[mask2]
    eta_auto.loc[mask1] = eta1.loc[mask1]
    rem_auto.loc[mask1] = pred1.loc[mask1]

    done_mask = df[TARGET_FINAL_FACT].notna()
    rem_auto = rem_auto.round()

    eta1 = eta1.mask(done_mask, pd.NaT)
    eta2 = eta2.mask(done_mask, pd.NaT)
    eta3 = eta3.mask(done_mask, pd.NaT)
    eta_auto = eta_auto.mask(done_mask, pd.NaT)
    rem_auto = rem_auto.mask(done_mask, np.nan)
    auto_stage = auto_stage.mask(done_mask, '')

    out = df.copy()
    out['ETA_прогноз_stage1'] = eta1
    out['ETA_прогноз_stage2'] = eta2
    out['ETA_прогноз_stage3'] = eta3
    out['ETA_прогноз_auto'] = eta_auto
    out['ETA_прогноз_auto_stage'] = auto_stage
    out['ETA_прогноз_auto_remaining_days'] = rem_auto

    out_pretty = _apply_output_aliases(out)

    log(f"\nСохранение: {output_path}")
    _save_excel_pretty(out_pretty, output_path)
    log("Готово.")

    return metrics
