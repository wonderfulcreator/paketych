# -*- coding: utf-8 -*-
"""
SebestoimostApp — расчёт себестоимости контейнера (Китай → РФ)
GUI: Tkinter
Excel: openpyxl

Файлы рядом с приложением:
- cbm_map.json (справочник "размер → CBM на 200 шт")
"""

from __future__ import annotations

import json
import os
import re
import sys
import subprocess
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, getcontext
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    # Drag&Drop (опционально)
    from tkinterdnd2 import DND_FILES, TkinterDnD  # type: ignore
    HAS_DND = True
except Exception:
    HAS_DND = False

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

APP_VERSION = "v0.4.0"

# Точность Decimal (хватит с запасом)
getcontext().prec = 28

# ---------------------------------------------------------------------
# Настройки / справочники
# ---------------------------------------------------------------------

# ТО(СБОР) — сумма зависит от кумулятивной таможенной стоимости (руб).
# Правила из твоей таблицы.
TO_SBOR_BRACKETS: List[Tuple[Decimal, Decimal]] = [
    (Decimal("200000"), Decimal("1067")),
    (Decimal("450000"), Decimal("2134")),
    (Decimal("1200000"), Decimal("4269")),
    (Decimal("2700000"), Decimal("11746")),
    (Decimal("4200000"), Decimal("16524")),
    (Decimal("5500000"), Decimal("21344")),
    (Decimal("7000000"), Decimal("27540")),
    (Decimal("999999999999"), Decimal("30000")),
]

DEFAULT_AFTER_RATE = Decimal("0.7")   # "После"
DEFAULT_BEFORE_RATE = Decimal("0.3")  # "До"

# Где на листе "Себестоимость" храним служебные параметры (скрытые колонки).
PARAM_LABEL_COL = "AA"
PARAM_VALUE_COL = "AB"
PARAM_START_ROW = 1

# В этой же зоне кладём рассчитанный ТО(СБОР) total
CELL_TO_SBOR_TOTAL = f"${PARAM_VALUE_COL}$13"

# ---------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------

_RE_NUM = re.compile(r"[-+]?\d[\d\s]*([.,]\d+)?")

def normalize_label(s: str) -> str:
    """Нормализация строк для поиска параметров (без транслитерации)."""
    s = (s or "").strip().upper()
    s = s.replace("Ё", "Е")
    s = re.sub(r"\s+", " ", s)
    return s

def parse_number(value) -> Optional[Decimal]:
    """Парсит число из ячейки: 1 234 567,89 / 1234567.89 / 'руб' и т.д."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except Exception:
            return None
    if isinstance(value, Decimal):
        return value

    s = str(value).strip()
    if not s:
        return None

    # Находим первое похоже-на-число
    m = _RE_NUM.search(s.replace("\u00a0", " "))  # NBSP → space
    if not m:
        return None

    num = m.group(0)
    # удаляем пробелы тысяч
    num = num.replace(" ", "")
    # если есть и ',' и '.', считаем ',' тысячным или наоборот? упрощённо:
    if "," in num and "." in num:
        # возьмём последнюю как десятичную
        if num.rfind(",") > num.rfind("."):
            num = num.replace(".", "")
            num = num.replace(",", ".")
        else:
            num = num.replace(",", "")
    else:
        num = num.replace(",", ".")

    try:
        return Decimal(num)
    except InvalidOperation:
        return None

def ensure_json_cbm_map(app_dir: Path) -> Path:
    """Убеждаемся, что cbm_map.json существует рядом с приложением."""
    p = app_dir / "cbm_map.json"
    if p.exists():
        return p
    # если нет — создаём минимальный дефолт
    default = {"MS": 0.0186, "M": 0.026208, "L": 0.0456, "XL": 0.056028}
    p.write_text(json.dumps(default, ensure_ascii=False, indent=2), encoding="utf-8")
    return p

def load_cbm_map(app_dir: Path) -> Dict[str, Decimal]:
    p = ensure_json_cbm_map(app_dir)
    data = json.loads(p.read_text(encoding="utf-8"))
    out: Dict[str, Decimal] = {}
    for k, v in data.items():
        try:
            out[str(k).strip()] = Decimal(str(v))
        except Exception:
            continue
    return out

def tokenize_name(name: str) -> List[str]:
    s = (name or "").upper().replace("Ё", "Е")
    # разделяем дефисы возле цифр/букв
    s = re.sub(r"(?<=\w)[-/](?=\w)", " ", s)
    # убираем лишнюю пунктуацию
    s = re.sub(r"[^\w\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.split() if s else []

def find_cbm_coeff(name: str, cbm_map: Dict[str, Decimal]) -> Optional[Decimal]:
    """
    Ищет размер в наименовании по токенам.
    Поддерживает multi-word ключи (например 'B 2XL', 'BC Premium').
    Приоритет — самый длинный ключ (по кол-ву токенов и длине).
    """
    tokens = tokenize_name(name)
    if not tokens:
        return None

    # подготовим ключи: токены + коэффициент
    keys: List[Tuple[List[str], Decimal, str]] = []
    for k, v in cbm_map.items():
        kt = tokenize_name(k)
        if kt:
            keys.append((kt, v, k))

    # сортировка: длиннее → раньше
    keys.sort(key=lambda x: (-len(x[0]), -len(x[2])))

    # поиск последовательности токенов
    for kt, val, _orig in keys:
        L = len(kt)
        for i in range(0, len(tokens) - L + 1):
            if tokens[i:i+L] == kt:
                return val
    return None

def calc_to_sbor_total(total_customs_rub: Decimal) -> Decimal:
    for limit, fee in TO_SBOR_BRACKETS:
        if total_customs_rub <= limit:
            return fee
    return TO_SBOR_BRACKETS[-1][1]

# ---------------------------------------------------------------------
# Чтение входного файла
# ---------------------------------------------------------------------

@dataclass
class InvoiceItem:
    name: str
    qty: Decimal
    amount_cny: Decimal
    cbm_coeff: Decimal  # на 200 шт

def detect_invoice_header(ws: Worksheet, max_rows: int = 60, max_cols: int = 35) -> Tuple[int, int, int, Optional[int], Optional[int]]:
    """
    Возвращает:
    (header_row, name_col, qty_col, amount_col, price_col)
    """
    name_kw = ["НАИМ", "ТОВАР", "ITEM", "DESC", "DESCRIPTION", "PRODUCT"]
    qty_kw = ["КОЛ", "QTY", "QUANTITY", "PCS", "ШТ"]
    amount_kw = ["AMOUNT", "СУММ", "TOTAL", "СТОИМ", "CNY"]
    price_kw = ["PRICE", "ЦЕН", "UNIT"]

    best = (0, 0, 0, None, None, 0)  # row, name, qty, amount, price, score

    for r in range(1, min(max_rows, ws.max_row) + 1):
        row_vals = []
        for c in range(1, min(max_cols, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                row_vals.append((c, normalize_label(v)))
        if not row_vals:
            continue

        name_c = next((c for c, s in row_vals if any(k in s for k in name_kw)), None)
        qty_c = next((c for c, s in row_vals if any(k in s for k in qty_kw)), None)
        amount_c = next((c for c, s in row_vals if any(k in s for k in amount_kw)), None)
        price_c = next((c for c, s in row_vals if any(k in s for k in price_kw)), None)

        score = 0
        score += 2 if name_c else 0
        score += 2 if qty_c else 0
        score += 1 if amount_c else 0
        score += 1 if price_c else 0

        if score > best[-1] and name_c and qty_c:
            best = (r, name_c, qty_c, amount_c, price_c, score)

    if best[-1] == 0:
        raise ValueError("Не удалось найти строку заголовков на листе инвойса. Проверь названия колонок.")

    return best[0], best[1], best[2], best[3], best[4]

def read_invoice(ws: Worksheet, cbm_map: Dict[str, Decimal]) -> List[InvoiceItem]:
    header_row, name_col, qty_col, amount_col, price_col = detect_invoice_header(ws)

    items: List[InvoiceItem] = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if name is None or str(name).strip() == "":
            # допускаем хвост с пустыми строками
            continue
        name_s = str(name).strip()

        qty = parse_number(ws.cell(r, qty_col).value)
        if qty is None:
            continue

        amount: Optional[Decimal] = None
        if amount_col:
            amount = parse_number(ws.cell(r, amount_col).value)
        if amount is None and price_col:
            price = parse_number(ws.cell(r, price_col).value)
            if price is not None:
                amount = qty * price

        if amount is None:
            raise ValueError(f"Не смог определить стоимость (CNY) для строки {r}: '{name_s}'")

        cbm_coeff = find_cbm_coeff(name_s, cbm_map)
        if cbm_coeff is None:
            raise ValueError(f"Не смог определить размер/CBM по наименованию: '{name_s}'. Добавь правило в cbm_map.json")

        items.append(InvoiceItem(name=name_s, qty=qty, amount_cny=amount, cbm_coeff=cbm_coeff))

    if not items:
        raise ValueError("В инвойсе не найдено ни одной позиции. Проверь структуру листа и строки.")
    return items

def read_params(ws: Worksheet) -> Dict[str, Decimal]:
    """
    Считывает параметры с 'Лист 2' в виде таблицы.
    Ищет текстовые метки по всему листу и рядом — числовые значения.
    """
    # Для каждого ключа: варианты, где каждый вариант — список подстрок (все должны присутствовать)
    patterns: Dict[str, List[List[str]]] = {
        "freight": [["ФРАХТ"], ["FREIGHT"]],
        "terminalka": [["ТЕРМИНАЛ"], ["ТЕРМИНАЛКА"], ["TERMINAL"]],
        "vyguzka": [["ВЫГРУЗ"], ["РАЗГРУЗ"], ["UNLOADING"]],
        "zhd": [["ЖД"], ["Ж/Д"], ["RAIL"]],
        "dop_port": [["ДОП", "ПОРТ"], ["ПОРТ", "РАСХОД"]],
        "dem_det": [["ДЕМ"], ["ДЕТ"], ["DEM", "DET"]],
        "dop_tam": [["ДОП", "ТАМОЖ"], ["ТАМОЖ", "РАСХОД"]],
        "vyvoz": [["ВЫВОЗ"], ["ДОСТАВК"]],
        "nepred": [["НЕПРЕД"], ["НЕПРЕДВИД"]],
        "rate_gtd": [["КУРС", "ГТД"], ["GTD", "RATE"], ["КОНВЕРТАЦ"]],
        "rate_pre": [["КУРС", "ПРЕДОПЛ"], ["PREPAY"], ["ПРЕДОПЛ"]],
        "rate_balance": [["КУРС", "БАЛАНС"], ["BALANCE"]],
    }

    required = ["freight", "terminalka", "vyguzka", "zhd", "dop_port", "dem_det", "dop_tam",
                "vyvoz", "nepred", "rate_gtd", "rate_pre", "rate_balance"]

    found: Dict[str, Decimal] = {}

    max_r = min(ws.max_row, 300)  # обычно хватает
    max_c = min(ws.max_column, 30)

    # соберём найденные подписи для отладки
    debug_labels: List[str] = []

    def match_key(label: str) -> Optional[str]:
        for key, variants in patterns.items():
            for parts in variants:
                ok = True
                for p in parts:
                    if p not in label:
                        ok = False
                        break
                if ok:
                    return key
        return None

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            lab = normalize_label(v)
            if len(lab) < 2:
                continue
            if c == 1 and len(debug_labels) < 60:
                debug_labels.append(lab)

            key = match_key(lab)
            if not key or key in found:
                continue

            # ищем число "рядом": сначала справа в этой строке, затем ниже
            candidate: Optional[Decimal] = None

            # справа 1..8 колонок
            for cc in range(c + 1, min(c + 9, max_c) + 1):
                candidate = parse_number(ws.cell(r, cc).value)
                if candidate is not None:
                    break

            # ниже (иногда значение под подписью)
            if candidate is None:
                for rr in range(r + 1, min(r + 3, max_r) + 1):
                    candidate = parse_number(ws.cell(rr, c + 1).value if (c + 1) <= max_c else None)
                    if candidate is not None:
                        break

            if candidate is not None:
                found[key] = candidate

    missing = [k for k in required if k not in found]
    if missing:
        # Плюс — попробуем самый частый формат "в колонке A подпись, в B значение"
        # (помогает, если общий поиск что-то не увидел из-за merged/формата)
        for r in range(1, max_r + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if not isinstance(a, str):
                continue
            lab = normalize_label(a)
            key = match_key(lab)
            if key and key not in found:
                val = parse_number(b)
                if val is not None:
                    found[key] = val

        missing = [k for k in required if k not in found]

    if missing:
        preview = ", ".join(debug_labels[:30]) if debug_labels else "(не удалось прочитать подписи)"
        raise ValueError(
            "Не найдены обязательные параметры на 'Лист 2': "
            + ", ".join(missing)
            + "\n\nПервые найденные подписи (A-колонка, нормализовано):\n"
            + preview
        )

    return found

# ---------------------------------------------------------------------
# Генерация выходного Excel
# ---------------------------------------------------------------------

THEMES = {
    "Clean Light": {
        "header_fill": "1F4E79",
        "header_font": "FFFFFF",
        "band_fill": "F3F6FA",
        "total_fill": "D9E2F3",
        "accent": "1F4E79",
    },
    "Minimal Gray": {
        "header_fill": "2F2F2F",
        "header_font": "FFFFFF",
        "band_fill": "F7F7F7",
        "total_fill": "E6E6E6",
        "accent": "2F2F2F",
    },
    "Green Accounting": {
        "header_fill": "1D6F42",
        "header_font": "FFFFFF",
        "band_fill": "F2F8F4",
        "total_fill": "DCEFE3",
        "accent": "1D6F42",
    },
}

def apply_theme(ws: Worksheet, n_rows: int, theme_name: str) -> None:
    theme = THEMES.get(theme_name, THEMES["Clean Light"])

    header_fill = PatternFill("solid", fgColor=theme["header_fill"])
    header_font = Font(bold=True, color=theme["header_font"])
    band_fill = PatternFill("solid", fgColor=theme["band_fill"])
    total_fill = PatternFill("solid", fgColor=theme["total_fill"])

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header row
    for c in range(1, 25):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 36

    # body rows banded
    for r in range(2, n_rows + 1):
        for c in range(1, 25):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c != 1 else "left", vertical="center", wrap_text=(c == 1))
            if r % 2 == 0:
                cell.fill = band_fill

    # total row (последняя строка)
    tr = n_rows
    for c in range(1, 25):
        cell = ws.cell(tr, c)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = border

    # widths
    widths = {
        1: 42, 2: 12, 3: 14, 4: 18, 5: 13, 6: 11, 7: 12, 8: 13, 9: 13, 10: 11,
        11: 9, 12: 16, 13: 10, 14: 16, 15: 7, 16: 7, 17: 16, 18: 18, 19: 18,
        20: 12, 21: 12, 22: 12, 23: 18, 24: 18
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:X{n_rows}"

def set_number_formats(ws: Worksheet, first_row: int, last_row: int) -> None:
    # B qty int
    for r in range(first_row, last_row + 1):
        ws[f"B{r}"].number_format = "#,##0"
        ws[f"C{r}"].number_format = "#,##0.00"
        for col in "DEFGHIJKLMNQR":
            ws[f"{col}{r}"].number_format = "#,##0.00"
        ws[f"O{r}"].number_format = "0.0%"
        ws[f"P{r}"].number_format = "0.0%"
        ws[f"S{r}"].number_format = "#,##0.00"
        ws[f"T{r}"].number_format = "0.0000"
        ws[f"U{r}"].number_format = "0.0000"
        ws[f"V{r}"].number_format = "0.0000"
        ws[f"W{r}"].number_format = "#,##0.00"
        ws[f"X{r}"].number_format = "#,##0.00"

def write_hidden_params(ws: Worksheet, params: Dict[str, Decimal], to_sbor_total: Decimal) -> None:
    labels = [
        ("freight", "Фрахт"),
        ("terminalka", "Терминалка"),
        ("vyguzka", "Выгрузка"),
        ("zhd", "ЖД"),
        ("dop_port", "Доп расходы в порту"),
        ("dem_det", "Дем/Дет"),
        ("dop_tam", "Доп расходы таможня"),
        ("vyvoz", "Вывоз (без НДС)"),
        ("nepred", "Непредвиденные затраты"),
        ("rate_gtd", "Курс ГТД"),
        ("rate_pre", "Курс предоплаты"),
        ("rate_balance", "Курс баланса"),
        ("to_sbor_total", "ТО(СБОР) total"),
    ]
    start = PARAM_START_ROW
    for i, (k, title) in enumerate(labels, start=0):
        r = start + i
        ws[f"{PARAM_LABEL_COL}{r}"] = title
        if k == "to_sbor_total":
            ws[f"{PARAM_VALUE_COL}{r}"] = float(to_sbor_total)
        else:
            ws[f"{PARAM_VALUE_COL}{r}"] = float(params[k])

    # hide columns
    ws.column_dimensions[PARAM_LABEL_COL].hidden = True
    ws.column_dimensions[PARAM_VALUE_COL].hidden = True

def generate_costing_xlsx(
    input_path: Path,
    output_path: Path,
    invoice_sheet: str,
    params_sheet: str,
    theme_name: str,
    cbm_map: Dict[str, Decimal],
) -> None:
    wb = openpyxl.load_workbook(input_path, data_only=False)
    if invoice_sheet not in wb.sheetnames:
        raise ValueError(f"Не найден лист инвойса '{invoice_sheet}'. Доступно: {', '.join(wb.sheetnames)}")
    if params_sheet not in wb.sheetnames:
        raise ValueError(f"Не найден лист параметров '{params_sheet}'. Доступно: {', '.join(wb.sheetnames)}")

    ws_inv = wb[invoice_sheet]
    ws_par = wb[params_sheet]

    params = read_params(ws_par)
    items = read_invoice(ws_inv, cbm_map)

    # Расчёт total TO(SBOR) в python (чтобы в итоговом файле не было таблицы диапазонов)
    freight = params["freight"]
    rate_gtd = params["rate_gtd"]

    # CBM по позициям
    v_list: List[Decimal] = [(it.qty / Decimal("200")) * it.cbm_coeff for it in items]
    v_sum = sum(v_list) if v_list else Decimal("0")
    if v_sum == 0:
        raise ValueError("Суммарная кубатура (CBM) равна 0 — проверь справочник cbm_map.json и количества.")

    # D = C * rate_gtd + I, где I = freight * U, U = V / sum(V)
    total_customs = Decimal("0")
    for it, v in zip(items, v_list):
        u = v / v_sum
        i_alloc = freight * u
        d = it.amount_cny * rate_gtd + i_alloc
        total_customs += d

    to_sbor_total = calc_to_sbor_total(total_customs)

    # Создаём выходной wb
    out = Workbook()
    ws = out.active
    ws.title = "Себестоимость"

    headers = [
        "Наименование товара", "Кол-во, шт", "Стоимость, CNY", "Таможенная стоимость, руб",
        "ТО (пошлина)", "ТО (СБОР)", "ТО (НДС)", "Терминалка", "Фрахт", "Выгрузка", "ЖД",
        "Доп расходы в порту", "Дем/Дет", "Доп расходы таможня", "После", "До",
        "Вывоз (без НДС)", "Непредвиденные затраты", "Затраты на конвертацию",
        "Стоимость КЭФ", "Кубатура КЭФ", "Кубатура, CBM",
        "С/С цена за кг, без НДС", "С/С цена за кг, с НДС"
    ]
    ws.append(headers)

    first_item_row = 2
    for idx, it in enumerate(items, start=0):
        r = first_item_row + idx
        ws[f"A{r}"] = it.name
        ws[f"B{r}"] = int(it.qty) if it.qty == it.qty.to_integral_value() else float(it.qty)
        ws[f"C{r}"] = float(it.amount_cny)

        # Формулы — сохраняем логику из примера, но ссылки на параметры заменяем на скрытые ячейки AB*
        ws[f"D{r}"] = f"=C{r}*$AB$10+I{r}"
        ws[f"E{r}"] = f"=D{r}*0.05"
        ws[f"F{r}"] = f"={CELL_TO_SBOR_TOTAL}*T{r}"
        ws[f"G{r}"] = f"=(D{r}+E{r})*1.22"

        ws[f"H{r}"] = f"=$AB$2*U{r}"
        ws[f"I{r}"] = f"=$AB$1*U{r}"
        ws[f"J{r}"] = f"=$AB$3*U{r}"
        ws[f"K{r}"] = f"=$AB$4*U{r}"
        ws[f"L{r}"] = f"=$AB$5*U{r}"
        ws[f"M{r}"] = f"=$AB$6*U{r}"
        ws[f"N{r}"] = f"=$AB$7*U{r}"

        ws[f"O{r}"] = float(DEFAULT_AFTER_RATE)
        ws[f"P{r}"] = float(DEFAULT_BEFORE_RATE)

        ws[f"Q{r}"] = f"=$AB$8*U{r}"
        ws[f"R{r}"] = f"=$AB$9*U{r}"
        ws[f"S{r}"] = f"=(C{r}*P{r}*$AB$11+C{r}*O{r}*$AB$12)*0.01"

        # Эти коэффициенты по шаблону — доля от итога
        # Важно: total_row ниже (после добавления строки Итого)
        # Поставим временные формулы, затем обновим после известного total_row
        ws[f"T{r}"] = ""  # будет после total_row
        ws[f"U{r}"] = ""  # будет после total_row

        # Кубатура CBM
        ws[f"V{r}"] = f"=B{r}/200*{float(it.cbm_coeff)}"

        ws[f"W{r}"] = (
            f"=(C{r}*P{r}*$AB$11+C{r}*O{r}*$AB$12+E{r}+F{r}+H{r}+I{r}+J{r}+K{r}"
            f"+L{r}+M{r}+N{r}+Q{r}+R{r}+S{r})/B{r}"
        )
        ws[f"X{r}"] = f"=1.22*W{r}"

    total_row = first_item_row + len(items)
    # Итого
    ws[f"A{total_row}"] = "Итого"
    ws[f"B{total_row}"] = f"=SUM(B{first_item_row}:B{total_row-1})"
    ws[f"C{total_row}"] = f"=SUM(C{first_item_row}:C{total_row-1})"
    ws[f"D{total_row}"] = f"=SUM(D{first_item_row}:D{total_row-1})"
    ws[f"E{total_row}"] = f"=SUM(E{first_item_row}:E{total_row-1})"
    ws[f"F{total_row}"] = f"=SUM(F{first_item_row}:F{total_row-1})"
    ws[f"G{total_row}"] = f"=SUM(G{first_item_row}:G{total_row-1})"
    ws[f"H{total_row}"] = f"=SUM(H{first_item_row}:H{total_row-1})"
    ws[f"I{total_row}"] = f"=SUM(I{first_item_row}:I{total_row-1})"
    ws[f"J{total_row}"] = f"=SUM(J{first_item_row}:J{total_row-1})"
    ws[f"K{total_row}"] = f"=SUM(K{first_item_row}:K{total_row-1})"
    ws[f"L{total_row}"] = f"=SUM(L{first_item_row}:L{total_row-1})"
    ws[f"M{total_row}"] = f"=SUM(M{first_item_row}:M{total_row-1})"
    ws[f"N{total_row}"] = f"=SUM(N{first_item_row}:N{total_row-1})"
    ws[f"Q{total_row}"] = f"=SUM(Q{first_item_row}:Q{total_row-1})"
    ws[f"R{total_row}"] = f"=SUM(R{first_item_row}:R{total_row-1})"
    ws[f"S{total_row}"] = f"=SUM(S{first_item_row}:S{total_row-1})"
    ws[f"V{total_row}"] = f"=SUM(V{first_item_row}:V{total_row-1})"

    # Теперь знаем total_row → проставляем T и U
    for idx in range(len(items)):
        r = first_item_row + idx
        ws[f"T{r}"] = f"=C{r}/$C${total_row}"
        ws[f"U{r}"] = f"=V{r}/$V${total_row}"

    # Скрытые параметры
    write_hidden_params(ws, params, to_sbor_total)

    # Оформление
    apply_theme(ws, total_row, theme_name)
    set_number_formats(ws, 2, total_row)

    # Excel Table (красивые фильтры/шапка)
    try:
        tab = Table(displayName="SebestoimostTable", ref=f"A1:X{total_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        # если конфликт имени таблицы — просто пропускаем
        pass

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out.save(output_path)

# ---------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------

class App:
    def __init__(self) -> None:
        if getattr(sys, "frozen", False):
            self.app_dir = Path(sys.executable).resolve().parent
        else:
            self.app_dir = Path(__file__).resolve().parent
        self.cbm_map = load_cbm_map(self.app_dir)

        if HAS_DND:
            self.root = TkinterDnD.Tk()
        else:
            self.root = tk.Tk()

        self.root.title(f"SebestoimostApp — расчёт себестоимости ({APP_VERSION})")
        self.root.geometry("980x560")

        self.in_path = tk.StringVar()
        self.out_path = tk.StringVar()
        self.invoice_sheet = tk.StringVar(value="Лист 1")
        self.params_sheet = tk.StringVar(value="Лист 2")
        self.theme = tk.StringVar(value="Clean Light")

        self._build_ui()

        if HAS_DND:
            self._enable_dnd()

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)

        title = ttk.Label(main, text="Расчёт себестоимости контейнера", font=("Segoe UI", 16, "bold"))
        title.grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 10))

        # Input
        ttk.Label(main, text="Входной XLSX:").grid(row=1, column=0, sticky="w")
        e_in = ttk.Entry(main, textvariable=self.in_path, width=92)
        e_in.grid(row=1, column=1, columnspan=3, sticky="we", padx=(8, 8))
        ttk.Button(main, text="Выбрать…", command=self.pick_input).grid(row=1, column=4, sticky="we")
        ttk.Button(main, text="Очистить", command=lambda: self.in_path.set("")).grid(row=1, column=5, sticky="we", padx=(8, 0))

        # Output
        ttk.Label(main, text="Выходной XLSX:").grid(row=2, column=0, sticky="w", pady=(6, 0))
        e_out = ttk.Entry(main, textvariable=self.out_path, width=92)
        e_out.grid(row=2, column=1, columnspan=3, sticky="we", padx=(8, 8), pady=(6, 0))
        ttk.Button(main, text="Куда сохранить…", command=self.pick_output).grid(row=2, column=4, sticky="we", pady=(6, 0))
        ttk.Button(main, text="Авто-имя", command=self.auto_name).grid(row=2, column=5, sticky="we", padx=(8, 0), pady=(6, 0))

        # Settings
        box = ttk.LabelFrame(main, text="Настройки", padding=10)
        box.grid(row=3, column=0, columnspan=6, sticky="we", pady=(14, 0))
        ttk.Label(box, text="Лист инвойса:").grid(row=0, column=0, sticky="w")
        ttk.Entry(box, textvariable=self.invoice_sheet, width=18).grid(row=0, column=1, sticky="w", padx=(8, 24))
        ttk.Label(box, text="Лист параметров:").grid(row=0, column=2, sticky="w")
        ttk.Entry(box, textvariable=self.params_sheet, width=18).grid(row=0, column=3, sticky="w", padx=(8, 24))
        ttk.Label(box, text="Тема оформления:").grid(row=0, column=4, sticky="w")
        ttk.Combobox(box, textvariable=self.theme, values=list(THEMES.keys()), width=18, state="readonly").grid(row=0, column=5, sticky="w", padx=(8, 0))

        # Buttons
        btns = ttk.Frame(main)
        btns.grid(row=4, column=0, columnspan=6, sticky="w", pady=(14, 0))
        ttk.Button(btns, text="Рассчитать", command=self.run_calc).grid(row=0, column=0, sticky="w")
        ttk.Button(btns, text="Открыть папку CBM", command=self.open_cbm_folder).grid(row=0, column=1, sticky="w", padx=(8, 0))
        ttk.Button(btns, text="Открыть выходной файл", command=self.open_output_file).grid(row=0, column=2, sticky="w", padx=(8, 0))

        # DnD hint
        hint = "Drag&Drop: включён" if HAS_DND else "Drag&Drop: установите tkinterdnd2 (опционально)"
        ttk.Label(main, text=hint).grid(row=5, column=0, columnspan=6, sticky="w", pady=(8, 0))

        # Log
        self.log = tk.Text(main, height=18, wrap="word")
        self.log.grid(row=6, column=0, columnspan=6, sticky="nsew", pady=(10, 0))
        main.rowconfigure(6, weight=1)
        main.columnconfigure(3, weight=1)

        self._log("Готово. Выберите файл и нажмите «Рассчитать».")

    def _enable_dnd(self) -> None:
        # Разрешаем drop на всё окно
        try:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind("<<Drop>>", self._on_drop)
        except Exception:
            pass

    def _on_drop(self, event) -> None:
        # event.data может содержать {path} или список
        data = str(event.data).strip()
        if data.startswith("{") and data.endswith("}"):
            data = data[1:-1]
        # если несколько — берём первый
        if " " in data and os.path.exists(data.split(" ")[0]):
            data = data.split(" ")[0]
        if data.lower().endswith(".xlsx"):
            self.in_path.set(data)
            self.auto_name()
            self._log(f"Drop: {data}")
        else:
            self._log("Drop: поддерживается только .xlsx")

    def _log(self, msg: str) -> None:
        self.log.insert("end", msg + "\n")
        self.log.see("end")

    def pick_input(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.in_path.set(p)
            self.auto_name()

    def pick_output(self) -> None:
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.out_path.set(p)

    def auto_name(self) -> None:
        p = self.in_path.get().strip()
        if not p:
            return
        in_path = Path(p)
        out = in_path.with_name(in_path.stem + "_CC.xlsx")
        self.out_path.set(str(out))

    def open_cbm_folder(self) -> None:
        p = ensure_json_cbm_map(self.app_dir)
        folder = p.parent
        try:
            os.startfile(str(folder))  # Windows
        except Exception:
            subprocess.Popen(["open", str(folder)]) if sys.platform == "darwin" else subprocess.Popen(["xdg-open", str(folder)])

    def open_output_file(self) -> None:
        p = self.out_path.get().strip()
        if not p or not Path(p).exists():
            messagebox.showinfo("Инфо", "Выходной файл ещё не создан или путь неверный.")
            return
        try:
            os.startfile(p)  # Windows
        except Exception:
            subprocess.Popen(["open", p]) if sys.platform == "darwin" else subprocess.Popen(["xdg-open", p])

    def run_calc(self) -> None:
        in_p = self.in_path.get().strip()
        out_p = self.out_path.get().strip()

        if not in_p or not Path(in_p).exists():
            messagebox.showerror("Ошибка", "Выберите существующий входной .xlsx файл.")
            return
        if not out_p:
            self.auto_name()
            out_p = self.out_path.get().strip()

        try:
            self._log("Старт расчёта…")
            generate_costing_xlsx(
                input_path=Path(in_p),
                output_path=Path(out_p),
                invoice_sheet=self.invoice_sheet.get().strip() or "Лист 1",
                params_sheet=self.params_sheet.get().strip() or "Лист 2",
                theme_name=self.theme.get().strip() or "Clean Light",
                cbm_map=self.cbm_map,
            )
            self._log(f"OK: сохранено → {out_p}")
            messagebox.showinfo("Готово", f"Файл сформирован:\n{out_p}")
        except Exception as e:
            self._log("ОШИБКА: " + str(e))
            messagebox.showerror("Ошибка", str(e))

def main() -> None:
    app = App()
    app.root.mainloop()

if __name__ == "__main__":
    main()
